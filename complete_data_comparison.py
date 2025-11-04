"""
Complete Data Comparison: CSV vs Excel (Power BI Reports)
Shows actual values from both sources with COLOR CODING for differences
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import warnings
warnings.filterwarnings('ignore')

# Files to compare
TABLES = {
    'SPEND_BY_CODE': {
        'csv_file': 'spend by code.csv',
        'excel_file': 'power bi actual report/Spend by code.xlsx',
        'description': 'Spend by Code Analysis'
    },
    'SPEND_BY_PRODUCT_TYPE': {
        'csv_file': 'Spend by product type.csv',
        'excel_file': 'power bi actual report/Spend by product type.xlsx',
        'description': 'Spend by Product Type Analysis'
    },
    'SPEND_BY_BILL_TYPE': {
        'csv_file': 'Spend by  bill type.csv',
        'excel_file': 'power bi actual report/Spend by  bill type.xlsx',
        'description': 'Spend by Bill Type Analysis'
    }
}

def clean_value(value):
    """Clean and normalize values for comparison"""
    if pd.isna(value) or value is None or value == '':
        return None
    
    if isinstance(value, str):
        value = value.strip()
        # Remove $, commas, %
        value = value.replace('$', '').replace(',', '').replace('%', '')
        
        try:
            return float(value)  # Always convert to float for consistency
        except:
            return value
    
    if isinstance(value, (int, float)):
        return float(value)  # Convert to float
    
    return value

def get_csv_data(csv_file):
    """Get all data from CSV file"""
    print(f"\n[CSV] Loading file: {csv_file}")
    
    try:
        csv_path = Path(csv_file)
        
        if not csv_path.exists():
            print(f"   Error: File not found")
            return None
        
        # Read CSV file
        df = pd.read_csv(csv_file)
        
        # Clean column names
        df.columns = [str(col).strip() for col in df.columns]
        
        # Remove blank rows
        df = df.dropna(how='all')
        
        # Remove filter rows
        first_col = df.columns[0]
        df = df[~df[first_col].astype(str).str.contains('Applied filters', case=False, na=False)]
        
        df = df.reset_index(drop=True)
        
        print(f"   Retrieved {len(df)} rows, {len(df.columns)} columns")
        
        return df
        
    except Exception as e:
        print(f"   Error: {str(e)}")
        return None

def get_excel_data(excel_file):
    """Get all data from Excel file"""
    print(f"\n[EXCEL] Loading file: {excel_file}")
    
    try:
        excel_path = Path(excel_file)
        
        if not excel_path.exists():
            print(f"   Error: File not found")
            return None
        
        # Read Excel file
        df = pd.read_excel(excel_file, engine='openpyxl', sheet_name=0)
        
        # Clean column names
        df.columns = [str(col).strip() for col in df.columns]
        
        # Remove blank rows
        df = df.dropna(how='all')
        
        # Remove filter rows
        first_col = df.columns[0]
        df = df[~df[first_col].astype(str).str.contains('Applied filters', case=False, na=False)]
        
        df = df.reset_index(drop=True)
        
        print(f"   Retrieved {len(df)} rows, {len(df.columns)} columns")
        
        return df
        
    except Exception as e:
        print(f"   Error: {str(e)}")
        return None

def compare_values(csv_val, excel_val, tolerance=0.01):
    """Compare two values with decimal tolerance"""
    # Handle None/NaN
    if pd.isna(csv_val) and pd.isna(excel_val):
        return False
    if pd.isna(csv_val) or pd.isna(excel_val):
        return True
    
    # Clean both values
    csv_clean = clean_value(csv_val)
    excel_clean = clean_value(excel_val)
    
    # Compare
    if isinstance(csv_clean, (int, float)) and isinstance(excel_clean, (int, float)):
        # Numeric comparison with tolerance (0.01 = ignore differences < 1 cent)
        diff = abs(csv_clean - excel_clean)
        return diff > tolerance
    else:
        # String comparison
        return str(csv_clean).strip() != str(excel_clean).strip()

def apply_color_coding(workbook, sheet_name, diff_cells):
    """Apply color coding to highlight differences with clear visual indicators"""
    # Define colors
    yellow_highlight = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Bright yellow for differences
    light_yellow = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Light yellow for status
    
    # Define fonts
    bold_red_font = Font(bold=True, color='CC0000')  # Bold red for status text
    
    ws = workbook[sheet_name]
    
    for cell_info in diff_cells:
        row = cell_info['row']
        csv_col = cell_info['csv_col']
        status_col = cell_info['status_col']
        excel_col = cell_info['excel_col']
        
        # Highlight CSV value (yellow background)
        csv_cell = ws.cell(row=row, column=csv_col)
        csv_cell.fill = yellow_highlight
        
        # Highlight STATUS cell (light yellow background, bold red text)
        status_cell = ws.cell(row=row, column=status_col)
        status_cell.fill = light_yellow
        status_cell.font = bold_red_font
        
        # Highlight Excel value (yellow background)
        excel_cell = ws.cell(row=row, column=excel_col)
        excel_cell.fill = yellow_highlight

def create_comparison_report():
    """Create comprehensive comparison report with color coding"""
    print("="*100)
    print("COMPLETE DATA COMPARISON: CSV vs EXCEL (Power BI Reports)")
    print("WITH COLOR CODING FOR DIFFERENCES")
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*100)
    
    output_dir = Path('validation_reports')
    output_dir.mkdir(exist_ok=True)
    
    # Clean up old CSV comparison files
    print("\n[CLEANUP] Removing old CSV comparison reports...")
    old_files = list(output_dir.glob('CSV_vs_Excel_Comparison_ColorCoded_*.xlsx'))
    for old_file in old_files:
        try:
            old_file.unlink()
            print(f"   Deleted: {old_file.name}")
        except Exception as e:
            print(f"   Warning: Could not delete {old_file.name}: {e}")
    
    if old_files:
        print(f"   Cleaned up {len(old_files)} old comparison(s)")
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = output_dir / f'CSV_vs_Excel_Comparison_ColorCoded_{timestamp}.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        summary_data = []
        all_diff_info = {}
        
        for table_name, config in TABLES.items():
            print(f"\n{'='*100}")
            print(f"PROCESSING: {table_name}")
            print(f"{'='*100}")
            
            # Get CSV data
            csv_df = get_csv_data(config['csv_file'])
            
            # Get Excel data
            excel_df = get_excel_data(config['excel_file'])
            
            if csv_df is None or excel_df is None:
                print(f"   Skipping {table_name} due to data load error")
                continue
            
            # Find differences
            total_cells = 0
            different_cells = 0
            diff_cells = []
            
            max_rows = min(len(csv_df), len(excel_df))
            
            # Create clear 3-column comparison: CSV | Status | Excel
            comparison_data = []
            
            for idx in range(max_rows):
                row_data = {'Row_Number': idx + 1}
                
                for col_idx, col in enumerate(csv_df.columns):
                    csv_val = csv_df.iloc[idx][col] if col in csv_df.columns else None
                    excel_val = excel_df.iloc[idx][col] if col in excel_df.columns else None
                    
                    total_cells += 1
                    
                    # Check if different (with 0.01 tolerance for decimal differences)
                    is_different = compare_values(csv_val, excel_val, tolerance=0.01)
                    
                    if is_different:
                        different_cells += 1
                        status = '❌ DIFFERENT'
                        diff_cells.append({
                            'row': idx + 2,  # +2 because of header and 0-indexing
                            'csv_col': col_idx * 3 + 2,  # CSV column position (Row + CSV cols)
                            'status_col': col_idx * 3 + 3,  # Status column position
                            'excel_col': col_idx * 3 + 4,  # Excel column position
                            'column_name': col,
                            'csv_value': csv_val,
                            'excel_value': excel_val
                        })
                    else:
                        status = '✓ Match'
                    
                    # Add values in clear 3-column format
                    row_data[f'{col}_CSV'] = csv_val
                    row_data[f'{col}_STATUS'] = status
                    row_data[f'{col}_Excel'] = excel_val
                
                comparison_data.append(row_data)
            
            # Create comparison DataFrame
            comparison_df = pd.DataFrame(comparison_data)
            
            # Write to Excel
            sheet_name = f'{table_name}'[:31]
            comparison_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Store diff info for color coding
            all_diff_info[sheet_name] = diff_cells
            
            # Summary
            match_percentage = ((total_cells - different_cells) / total_cells * 100) if total_cells > 0 else 100
            
            summary_data.append({
                'Table': table_name,
                'CSV_Rows': len(csv_df),
                'Excel_Rows': len(excel_df),
                'Total_Cells_Compared': total_cells,
                'Different_Cells': different_cells,
                'Matching_Cells': total_cells - different_cells,
                'Match_Percentage': f'{match_percentage:.2f}%'
            })
            
            print(f"\n   COMPARISON RESULTS:")
            print(f"   Total cells compared: {total_cells}")
            print(f"   Different cells: {different_cells}")
            print(f"   Match percentage: {match_percentage:.2f}%")
            print(f"   Color coding applied to {len(diff_cells)} differences")
        
        # Create summary sheet
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            print(f"\n   Created summary sheet")
    
    # Apply color coding
    print(f"\n{'='*100}")
    print("APPLYING COLOR CODING TO DIFFERENCES...")
    print(f"{'='*100}")
    
    workbook = load_workbook(output_file)
    
    for sheet_name, diff_cells in all_diff_info.items():
        if diff_cells:
            print(f"\n   Coloring {len(diff_cells)} differences in {sheet_name}")
            apply_color_coding(workbook, sheet_name, diff_cells)
    
    # Add legend sheet with clear explanation
    legend_data = pd.DataFrame([
        {'Format': 'Column Header', 'Explanation': 'Each data column has 3 sub-columns: CSV | STATUS | Excel'},
        {'Format': 'YELLOW Highlight', 'Explanation': 'Values are DIFFERENT between CSV and Excel'},
        {'Format': '❌ DIFFERENT', 'Explanation': 'Status indicator showing mismatch'},
        {'Format': '✓ Match', 'Explanation': 'Values are identical (no highlighting)'},
        {'Format': '', 'Explanation': ''},
        {'Format': 'How to Read', 'Explanation': 'Compare CSV value ← → Excel value side-by-side'},
        {'Format': 'Yellow Row', 'Explanation': 'These values DO NOT match - review both values'},
        {'Format': 'No Color', 'Explanation': 'These values MATCH perfectly'}
    ])
    
    ws = workbook.create_sheet('Legend')
    
    # Write header
    ws.cell(row=1, column=1, value='FORMAT').font = Font(bold=True)
    ws.cell(row=1, column=2, value='EXPLANATION').font = Font(bold=True)
    
    # Write data
    for r_idx, row in enumerate(legend_data.values, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Apply color coding to legend examples
    yellow_highlight = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    light_yellow = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    bold_red_font = Font(bold=True, color='CC0000')
    
    ws.cell(row=3, column=1).fill = yellow_highlight  # YELLOW Highlight example
    ws.cell(row=4, column=1).fill = light_yellow  # Status cell example
    ws.cell(row=4, column=1).font = bold_red_font
    ws.cell(row=8, column=1).fill = yellow_highlight  # Yellow Row example
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 70
    
    workbook.save(output_file)
    
    print(f"\n{'='*100}")
    print(f"REPORT GENERATED SUCCESSFULLY WITH CLEAR COMPARISON FORMAT")
    print(f"{'='*100}")
    print(f"\nFile: {output_file}")
    print(f"\n📊 REPORT FORMAT:")
    print(f"  Each column is split into 3 parts: CSV | STATUS | Excel")
    print(f"\n🎨 COLOR CODING:")
    print(f"  YELLOW HIGHLIGHT = Values are DIFFERENT (review both CSV and Excel values)")
    print(f"  ❌ DIFFERENT = Status indicator showing mismatch")
    print(f"  ✓ Match = Values are identical (no highlighting)")
    print(f"  No Color = Perfect match between CSV and Excel")
    print(f"\n💡 HOW TO USE:")
    print(f"  1. Yellow rows show differences - compare CSV vs Excel values side-by-side")
    print(f"  2. STATUS column clearly marks differences with ❌ DIFFERENT")
    print(f"  3. No highlighting = values match perfectly")
    print(f"  4. Check 'Legend' sheet for detailed explanation")
    print(f"\n📝 SHEETS INCLUDED:")
    print(f"  - Summary: Match percentages for all tables")
    for table_name in TABLES.keys():
        print(f"  - {table_name}: CSV ← → Excel comparison")
    print(f"  - Legend: Color coding and format explanation")
    print(f"\nDifferences are CLEARLY HIGHLIGHTED with yellow so you can instantly see mismatches!\n")

if __name__ == "__main__":
    create_comparison_report()
